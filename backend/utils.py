from datetime import timedelta, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
import sqlite3

import jwt
from passlib.context import CryptContext
from sqlalchemy.orm import Session

from models.models_class_info import Grade, Major, Dorm, Classes
from models.models_sanitation import Sanitation
from db.database import get_db
from core.config import settings

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)


def get_password_hash(password):
    return pwd_context.hash(password)


def create_jwt_token(data: dict):
    expiration = str(datetime.utcnow()) + str(settings.JWT_EXPIRE_MINUTES)
    to_encode = data.copy()
    to_encode.update({"exp": expiration})
    encoded_jwt = jwt.encode(to_encode, settings.SECRET_KEY, algorithm=settings.JWT_ALGORITHM)
    return encoded_jwt


def generate_excel(db: Session) -> BytesIO:
    wb = Workbook()

    # 查询GradeTable中所有的gradeName
    grades = db.query(Grade).all()

    # 删除默认创建的工作表
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 创建边框样式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # 创建居中对齐样式
    center_alignment = Alignment(horizontal='center', vertical='center')

    for grade in grades:
        grade_name = grade.GradeName
        ws = wb.create_sheet(title=str(grade_name))
        wb.active = ws

        # 初始化行索引
        row_index = 3

        # 查询与当前 gradeName 相关的所有专业
        majors = db.query(Major).filter(Major.GradeID == grade.GradeID).all()

        # 对于每一个专业，查询相关的班级
        for major in majors:
            classes = db.query(Classes).filter(Classes.MajorID == major.MajorID).all()

            for class_ in classes:
                # 查询与当前班级相关的所有宿舍
                dorms = db.query(Dorm).filter(Dorm.ClassID == class_.ClassID).all()

                # 合并单元格并设置班级名称
                if len(dorms) > 0:
                    ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index + len(dorms) - 1,
                                   end_column=1)
                    ws.cell(row=row_index, column=1).value = f"{grade_name}{major.MajorName}{class_.ClassName}"

                # 填充宿舍名称
                for i, dorm in enumerate(dorms):
                    ws.cell(row=row_index + i, column=2).value = dorm.DormName

                    # 查询与当前宿舍相关的所有卫生记录
                    sanitation_records = db.query(Sanitation).filter(Sanitation.DormID == dorm.DormID).all()
                    # 在开始填充卫生状况之前，初始化一个字典来存储优秀次数
                    excellent_counts = {}

                    # 填充卫生状况
                    for sanitation in sanitation_records:
                        week_number = sanitation.WeekNumber
                        if sanitation.Status == "优秀":
                            excellent_counts[week_number] = excellent_counts.get(week_number, 0) + 1

                        # 将 weekday_str 转换为数字索引，假设 "周一" 对应 0，"周二" 对应 1，以此类推
                        weekday_to_index = {1: 0, 2: 1, 3: 2, 4: 3, 5: 4}
                        weekday = weekday_to_index.get(sanitation.Weekday, -1)  # 如果找不到对应的索引，返回 -1

                        # 计算应填充到的单元格的列索引
                        col_index_for_week = 4 + (week_number - 1) * 6  # 从第4列开始，每周有6列（包括“优秀次数”列）
                        col_index_for_status = col_index_for_week + weekday  # 根据 Weekday 的数字索引来确定

                        # 获取单元格
                        cell = ws.cell(row=row_index + i, column=col_index_for_status)

                        # 填充状态
                        cell.value = sanitation.Status

                        # 根据状态设置单元格的底色
                        if sanitation.Status == "优秀":
                            cell.fill = green_fill
                        elif sanitation.Status == "不及格":
                            cell.fill = red_fill

                    # 在填充 Excel 表格的时候，使用 excellent_counts 字典来获取优秀次数
                    for week_number, count in excellent_counts.items():
                        col_index_for_excellent = 3 + (week_number - 1) * 6  # 从第3列开始，每周有6列（包括“优秀次数”列）
                        ws.cell(row=row_index + i, column=col_index_for_excellent).value = count

                    # 清空 excellent_counts 字典以便用于下一个宿舍
                    excellent_counts.clear()

                # 更新行索引
                row_index += len(dorms)

        # 合并单元格并设置值
        ws.merge_cells('A1:A2')
        ws['A1'].value = '班级'
        ws['A1'].border = thin_border
        ws['A1'].alignment = center_alignment

        ws.merge_cells('B1:B2')
        ws['B1'].value = '宿舍'
        ws['B1'].border = thin_border
        ws['B1'].alignment = center_alignment

        # 设置 C1 的值为 '优秀次数'
        ws.cell(row=1, column=3).value = '优秀次数'
        ws.cell(row=1, column=3).border = thin_border
        ws.cell(row=1, column=3).alignment = center_alignment

        # 合并 C1 和 C2
        ws.merge_cells('C1:C2')

        # 初始化列索引
        col_index = 4  # 从第3列开始，即C列

        # 循环处理每一周
        for week in range(1, 17):
            # 合并单元格
            ws.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index+4)
            ws.merge_cells(start_row=1, start_column=col_index-1, end_row=2, end_column=col_index-1)

            # 设置合并后单元格的值
            ws.cell(row=1, column=col_index).value = f'第{week}周'
            ws.cell(row=1, column=col_index-1).value = '优秀次数'

            # 设置 D2 到 H2 的值
            weekdays = ['周一', '周二', '周三', '周四', '周五']
            for i, weekday in enumerate(weekdays, start=0):
                cell = ws.cell(row=2, column=col_index+i)
                cell.value = weekday

            # 更新列索引
            col_index += 6

        # 设置所有相关单元格的边框样式和对齐方式
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=col_index - 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment

    # 使用BytesIO对象代替直接保存到文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


